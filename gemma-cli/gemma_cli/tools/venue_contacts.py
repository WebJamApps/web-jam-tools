"""Venue contacts tools: lookup, web-derive, update against Josh's xlsx.

Three tools wired through to Josh's canonical venue contacts spreadsheet at
`~/Dropbox/joshandmariamusic/Gig Booking Worksheet 2025.xlsx` (per CLAUDE.md
"XLSX TRACKING SPREADSHEET" — the Dropbox path is the authoritative master).

  - `lookup_venue_contact(venue_name)` — find an existing row in the xlsx by
    venue name. Returns contact details (booker name, email, phone, status) or
    a not-found marker.
  - `lookup_venue_email_on_web(website_url)` — fetch the venue's website
    (homepage + /contact + /booking + /about), extract email addresses with a
    regex, return the best candidate. Used when the xlsx lookup misses.
  - `update_venue_contact(venue_name, ...)` — write contact details back to
    the xlsx. Updates an existing row if found by name, otherwise appends a new
    row. Persists immediately to disk.

The xlsx has a multi-section layout: row 1 = header row ('Name', 'Contact',
'email', 'phone #', 'type of gig', 'last played', 'comments', ...) and then
sub-headers appear at section breaks ('Current Gigs', 'Venue', etc.). Lookup
iterates ALL rows and matches on column-1 substring (case-insensitive); the
section-header rows have short generic strings that won't collide with real
venue names.
"""

from __future__ import annotations

import re
from difflib import SequenceMatcher
from typing import Any
from urllib.parse import urljoin, urlparse

import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook

from gemma_cli.llm import Tool


XLSX_PATH = "/home/joshua/Dropbox/joshandmariamusic/Gig Booking Worksheet 2025.xlsx"

# Fuzzy-match threshold for venue-name lookup. 0.85 catches the Olde Salem
# Brewing / Olde Salem Brewery case (similarity ~0.95) and similar
# Brewery↔Brewing / Inn↔Tavern suffix swaps, without conflating clearly
# distinct venues. Tuned 2026-05-21 after the Olde Salem duplicate-row
# regression that inserted row 81 alongside the canonical row 6.
_FUZZY_MATCH_THRESHOLD = 0.85

# Common business-suffix words stripped from venue names before fuzzy
# comparison (added 2026-05-21). The Cavendish run inserted a duplicate row
# because "Cavendish Brewing Company" vs "Cavendish Brewing" only scored 0.81
# on SequenceMatcher — under the 0.85 threshold. Canonicalizing both sides
# to "cavendish" before the ratio makes them match exactly.
_BUSINESS_SUFFIXES = (
    "brewing", "brewery", "company", "co", "llc", "inc", "incorporated",
    "restaurant", "tavern", "pub", "bar", "grill", "cafe", "café", "bistro",
    "kitchen", "winery", "distillery", "ltd", "limited",
)

# Column indexes in the xlsx (1-based per openpyxl). Verified 2026-05-18 by
# reading the sheet headers in Sheet1 row 1: 'Name', 'Contact', 'email',
# 'phone #', 'type of gig', 'last played', 'comments', 'location', 'Status',
# 'Date called', 'Notes', 'Callback', 'Venue Name', 'Date verified', 'Phone'.
_COL_NAME = 1
_COL_CONTACT = 2
_COL_EMAIL = 3
_COL_PHONE = 4
_COL_TYPE_OF_GIG = 5
_COL_LAST_PLAYED = 6
_COL_COMMENTS = 7
_COL_LOCATION = 8
_COL_STATUS = 9
_COL_DATE_CALLED = 10
_COL_NOTES = 11
_COL_CALLBACK = 12

# Section-header rows whose col-1 values are generic words, not real venue
# names. Used to skip false matches during lookup.
_SECTION_HEADER_TOKENS = frozenset({
    "name", "venue", "current gigs", "venues to contact", "contact",
    "previously played", "future leads",
})

# Email regex — RFC-lite. Matches the common forms found on venue websites.
_EMAIL_RE = re.compile(r"[a-zA-Z0-9._%+\-]+@[a-zA-Z0-9.\-]+\.[a-zA-Z]{2,}")

# Common venue-website pages where booking emails tend to live. Tried in order.
# Restaurant/brewery venues frequently put their booking email on an /events
# or /private-events page rather than a generic /contact — added 2026-05-21
# after the 419 West miss (info@419-west.com lived only on /events).
_CONTACT_PATHS = (
    "/",
    "/contact", "/contact-us", "/contact/",
    "/booking", "/book", "/booking/",
    "/about", "/about-us",
    "/events", "/events/",
    "/private-events", "/private-events/",
    "/private-dining", "/private-dining/",
    "/parties", "/parties/",
    "/host", "/host-an-event",
    "/inquiry", "/inquiries",
)

# Emails we should never propose as a venue booker. Common webhost / privacy /
# generic-platform addresses that appear on many sites without being the venue.
_EMAIL_BLOCKLIST_DOMAINS = frozenset({
    "wixpress.com", "sentry.io", "godaddy.com", "wordpress.com", "squarespace.com",
    "gravatar.com", "example.com", "example.org",
})
_EMAIL_BLOCKLIST_LOCALPARTS = frozenset({
    "noreply", "no-reply", "donotreply", "do-not-reply", "postmaster", "abuse",
    "webmaster", "privacy", "security", "support@wix",
})


def _is_section_header(name_cell: Any) -> bool:
    """True if a row's col-1 value looks like a section header rather than a venue."""
    if not isinstance(name_cell, str):
        return False
    return name_cell.strip().lower() in _SECTION_HEADER_TOKENS


def _normalize(s: Any) -> str:
    return str(s).strip().lower() if s is not None else ""


def _canonicalize_venue_name(name: Any) -> str:
    """Lowercase + strip business-suffix words + punctuation for fuzzy match.

    Without this, "Cavendish Brewing" and "Cavendish Brewing Company" sit at
    SequenceMatcher ratio ~0.81 — below the 0.85 fuzzy threshold — and the
    second form would insert a duplicate row. Canonicalizing both sides to
    "cavendish" before the ratio makes them match exactly.

    Strips suffix words ONLY from the trailing position so a venue like
    "Brewery Burger Bar" (hypothetical) doesn't get its leading "Brewery"
    chopped. Iterates because a name can end in multiple suffixes
    (e.g. "X Brewing Company" → strip "Company" → "X Brewing" → strip
    "Brewing" → "X").
    """
    s = _normalize(name)
    # Punctuation that's noise for venue matching. Don't strip apostrophes
    # (e.g. "Bobby's") or ampersands inside the core (e.g. "Mac & Cheese"
    # would lose its meaning) — only the obvious separators.
    s = re.sub(r"[.,]", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    changed = True
    while changed:
        changed = False
        for suffix in _BUSINESS_SUFFIXES:
            # Only strip when there's at least one word before the suffix.
            if s.endswith(" " + suffix):
                s = s[: -(len(suffix) + 1)].strip()
                changed = True
                break
    return s


def _row_to_contact_dict(name: Any, row: tuple) -> dict[str, Any]:
    """Build the contact dict from an xlsx row. Used by both exact and fuzzy paths."""
    return {
        "found": True,
        "venue_name": str(name).strip(),
        "contact": _str_or_empty(row, _COL_CONTACT),
        "email": _str_or_empty(row, _COL_EMAIL),
        "phone": _str_or_empty(row, _COL_PHONE) or _str_or_empty(row, 15),  # col 15 = 'Phone' fallback
        "type_of_gig": _str_or_empty(row, _COL_TYPE_OF_GIG),
        "last_played": _str_or_empty(row, _COL_LAST_PLAYED),
        "comments": _str_or_empty(row, _COL_COMMENTS),
        "location": _str_or_empty(row, _COL_LOCATION),
        "status": _str_or_empty(row, _COL_STATUS),
        "notes": _str_or_empty(row, _COL_NOTES),
    }


def lookup_venue_contact(venue_name: str) -> dict[str, Any]:
    """Look up a venue in the xlsx by name.

    Two-pass strategy:
    1. Substring match (case-insensitive) — handles exact + obvious partials.
    2. Fuzzy match (difflib.SequenceMatcher, threshold 0.85) — catches near-
       spelling variants like "Olde Salem Brewing" ↔ "Olde Salem Brewery"
       that would otherwise produce a duplicate row on update.

    Return dict includes `match_type` ("exact" or "fuzzy") and, when fuzzy,
    `similarity` (0.0-1.0). Callers should treat fuzzy hits as candidates
    needing user confirmation, not auto-applied updates.
    """
    target = venue_name.strip().lower()
    if not target:
        return {"found": False, "error": "venue_name is required."}
    # Canonical form for the fuzzy + suffix-aware substring passes. See
    # _canonicalize_venue_name docstring (2026-05-21 Cavendish duplicate fix).
    target_canonical = _canonicalize_venue_name(venue_name)
    try:
        wb = load_workbook(XLSX_PATH, read_only=True, data_only=True)
    except Exception as exc:
        return {"found": False, "error": f"Could not open xlsx: {type(exc).__name__}: {exc}"}
    try:
        ws = wb.active
        best_fuzzy: tuple[float, Any, tuple] | None = None
        for row in ws.iter_rows(min_row=1, values_only=True):
            if not row:
                continue
            name = row[_COL_NAME - 1]
            if not name or _is_section_header(name):
                continue
            row_norm = _normalize(name)
            # Pass 1: raw substring (existing exact-ish behavior)
            if target in row_norm:
                result = _row_to_contact_dict(name, row)
                result["match_type"] = "exact"
                result["query"] = venue_name
                return result
            # Pass 1b: canonical substring (added 2026-05-21). After stripping
            # business suffixes from both sides, "cavendish brewing company"
            # collapses to "cavendish" which substring-matches "cavendish
            # brewing"'s canonical "cavendish". Treat this as exact too — it's
            # the same venue with a different suffix, not a fuzzy guess.
            row_canonical = _canonicalize_venue_name(name)
            if target_canonical and row_canonical and (
                target_canonical in row_canonical
                or row_canonical in target_canonical
            ):
                result = _row_to_contact_dict(name, row)
                result["match_type"] = "exact"
                result["query"] = venue_name
                return result
            # Pass 2 setup: fuzzy on canonical forms (suffix-insensitive). The
            # raw forms catch Brewing↔Brewery; canonical forms catch
            # "X" vs "X Brewing Company".
            similarity = SequenceMatcher(
                None, target_canonical or target, row_canonical or row_norm
            ).ratio()
            if similarity >= _FUZZY_MATCH_THRESHOLD and (
                best_fuzzy is None or similarity > best_fuzzy[0]
            ):
                best_fuzzy = (similarity, name, row)
        if best_fuzzy is not None:
            similarity, name, row = best_fuzzy
            result = _row_to_contact_dict(name, row)
            result["match_type"] = "fuzzy"
            result["similarity"] = round(similarity, 3)
            result["query"] = venue_name
            return result
        return {
            "found": False,
            "venue_name": venue_name,
            "match_type": "none",
            "query": venue_name,
            "note": (
                f"No row in the xlsx matched '{venue_name}' (exact or fuzzy "
                f"≥{_FUZZY_MATCH_THRESHOLD}). Next step: call "
                "`lookup_venue_email_on_web` with the venue's website URL, then "
                "(if a booker email is found) call `update_venue_contact` to "
                "add the venue to the xlsx for next time."
            ),
        }
    finally:
        wb.close()


def _str_or_empty(row: tuple, col_1based: int) -> str:
    """Return a stripped string for row[col-1], or empty string if None / out-of-bounds."""
    idx = col_1based - 1
    if idx < 0 or idx >= len(row):
        return ""
    val = row[idx]
    return str(val).strip() if val is not None else ""


def _is_useful_email(email: str, website_domain: str | None = None) -> bool:
    """Filter out obvious non-venue emails (webhost, noreply, etc.)."""
    email_lower = email.lower().strip()
    local, _, domain = email_lower.partition("@")
    if not domain:
        return False
    # Subdomain-aware blocklist (2026-05-21): the scraper found
    # `<hash>@sentry.wixpress.com` on castroanoke.com — the literal-match check
    # missed it because the blocklist had `wixpress.com` only. Treat any
    # subdomain of a blocked domain as blocked too.
    if any(domain == bad or domain.endswith("." + bad) for bad in _EMAIL_BLOCKLIST_DOMAINS):
        return False
    if local in _EMAIL_BLOCKLIST_LOCALPARTS:
        return False
    if any(local.startswith(bad) for bad in ("noreply", "no-reply", "donotreply")):
        return False
    # Machine-generated local parts (Sentry DSN public keys, tracking pixels,
    # transactional-email signing tokens) are 32+ chars of pure hex. Real
    # booker addresses are short and contain non-hex letters.
    if len(local) >= 32 and all(c in "0123456789abcdef" for c in local):
        return False
    return True


def lookup_venue_email_on_web(website_url: str) -> dict[str, Any]:
    """Fetch the venue's website and extract email addresses from common contact pages."""
    url = website_url.strip()
    if not url:
        return {"found": False, "error": "website_url is required."}
    if not url.startswith(("http://", "https://")):
        url = "https://" + url
    parsed = urlparse(url)
    base = f"{parsed.scheme}://{parsed.netloc}"
    domain = parsed.netloc.lower().lstrip("www.")

    headers = {
        "User-Agent": (
            "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 "
            "(KHTML, like Gecko) Chrome/120.0 Safari/537.36"
        ),
    }
    emails_found: dict[str, str] = {}  # email -> source URL

    for path in _CONTACT_PATHS:
        candidate = urljoin(base, path)
        try:
            resp = requests.get(candidate, headers=headers, timeout=10, allow_redirects=True)
        except Exception:
            continue
        if resp.status_code != 200:
            continue
        # Search the raw HTML body + any mailto: links via BeautifulSoup.
        html = resp.text
        for match in _EMAIL_RE.findall(html):
            if match not in emails_found and _is_useful_email(match, domain):
                emails_found[match] = candidate
        try:
            soup = BeautifulSoup(html, "html.parser")
            for a in soup.select('a[href^="mailto:"]'):
                href = a.get("href", "")
                addr = href.replace("mailto:", "").split("?")[0].strip()
                if addr and _is_useful_email(addr, domain) and addr not in emails_found:
                    emails_found[addr] = candidate
        except Exception:
            pass
        # Stop early ONLY if we've found a same-domain email — the venue's own
        # address is the canonical booker. If we've only found third-party
        # gmail/yahoo addresses (often dev/designer credits embedded in JS, or
        # personal addresses unrelated to bookings), keep scanning the rest of
        # _CONTACT_PATHS — a real venue email may live on /events or similar.
        # 2026-05-21: 419 West had impallari@gmail.com (typography designer)
        # in homepage JS; the old break here caused us to miss info@419-west.com
        # which was on /events.
        if any(addr.split("@", 1)[1].lower() == domain for addr in emails_found):
            break

    if not emails_found:
        return {
            "found": False,
            "website_url": url,
            "note": (
                f"No usable booker email found on {url} or its common contact "
                "pages. Ask Josh for the booker email directly."
            ),
        }
    # Prefer emails whose domain matches the venue's website domain (likeliest
    # real booker email) over generic gmail/yahoo addresses found on the page.
    sorted_emails = sorted(
        emails_found.items(),
        key=lambda kv: (0 if kv[0].split("@", 1)[1].lower() == domain else 1, kv[0]),
    )
    primary = sorted_emails[0]
    return {
        "found": True,
        "website_url": url,
        "best_email": primary[0],
        "source_page": primary[1],
        "all_emails_found": [{"email": e, "source": s} for e, s in sorted_emails],
    }


def update_venue_contact(
    venue_name: str,
    email: str = "",
    contact: str = "",
    phone: str = "",
    location: str = "",
    notes: str = "",
) -> dict[str, Any]:
    """Update or insert a venue contact row in the xlsx and save."""
    name = venue_name.strip()
    if not name:
        return {"updated": False, "error": "venue_name is required."}
    try:
        wb = load_workbook(XLSX_PATH)  # NOT read_only — we need to write.
    except Exception as exc:
        return {"updated": False, "error": f"Could not open xlsx: {type(exc).__name__}: {exc}"}
    try:
        ws = wb.active
        target = name.lower()
        existing_row = None
        for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
            if not row:
                continue
            cell_name = row[_COL_NAME - 1]
            if not cell_name or _is_section_header(cell_name):
                continue
            if target in _normalize(cell_name):
                existing_row = row_idx
                break

        if existing_row is not None:
            action = "updated"
            row_idx = existing_row
            if email:
                ws.cell(row=row_idx, column=_COL_EMAIL).value = email.strip()
            if contact:
                ws.cell(row=row_idx, column=_COL_CONTACT).value = contact.strip()
            if phone:
                ws.cell(row=row_idx, column=_COL_PHONE).value = phone.strip()
            if location:
                ws.cell(row=row_idx, column=_COL_LOCATION).value = location.strip()
            if notes:
                # Append rather than overwrite if notes already present.
                existing_notes = ws.cell(row=row_idx, column=_COL_NOTES).value
                merged = f"{existing_notes}; {notes.strip()}" if existing_notes else notes.strip()
                ws.cell(row=row_idx, column=_COL_NOTES).value = merged
        else:
            action = "created"
            row_idx = ws.max_row + 1
            ws.cell(row=row_idx, column=_COL_NAME).value = name
            if contact:
                ws.cell(row=row_idx, column=_COL_CONTACT).value = contact.strip()
            if email:
                ws.cell(row=row_idx, column=_COL_EMAIL).value = email.strip()
            if phone:
                ws.cell(row=row_idx, column=_COL_PHONE).value = phone.strip()
            if location:
                ws.cell(row=row_idx, column=_COL_LOCATION).value = location.strip()
            if notes:
                ws.cell(row=row_idx, column=_COL_NOTES).value = notes.strip()

        wb.save(XLSX_PATH)
        return {
            "updated": True,
            "action": action,
            "row": row_idx,
            "venue_name": name,
            "note": (
                f"Venue {action} in xlsx at row {row_idx}. The file at "
                f"{XLSX_PATH} is saved."
            ),
        }
    finally:
        wb.close()


TOOLS: list[Tool] = [
    Tool(
        name="lookup_venue_contact",
        description=(
            "Look up a venue's contact details (booker name, email, phone, status) "
            "in Josh's Gig Booking Worksheet xlsx. This is the FIRST step when you "
            "need a TO address for a venue-outreach email — check here before "
            "asking Josh or scraping the web. Returns {found: True, email: ..., "
            "...} if the venue is in the xlsx, otherwise {found: False, note: "
            "'next step is web lookup'}."
        ),
        parameters={
            "type": "object",
            "properties": {
                "venue_name": {
                    "type": "string",
                    "description": (
                        "The venue's name as it would appear in the spreadsheet "
                        "(e.g. 'Solstice Farm Brewery'). Case-insensitive "
                        "substring match — 'Solstice' alone would also match."
                    ),
                },
            },
            "required": ["venue_name"],
        },
        handler=lookup_venue_contact,
    ),
    Tool(
        name="lookup_venue_email_on_web",
        description=(
            "Fetch a venue's website and extract booker email addresses from the "
            "homepage and common contact pages (/contact, /booking, /about). Use "
            "this AFTER lookup_venue_contact returns found=False, when you have "
            "the venue's website URL from the task. Returns {found: True, "
            "best_email: 'info@venue.com', source_page: '...'} or {found: "
            "False, note: 'ask Josh'}. Filters out webhost / noreply / privacy "
            "addresses automatically."
        ),
        parameters={
            "type": "object",
            "properties": {
                "website_url": {
                    "type": "string",
                    "description": (
                        "The venue's website URL (with or without https://). "
                        "Comes from the task's VENUE DETAILS section."
                    ),
                },
            },
            "required": ["website_url"],
        },
        handler=lookup_venue_email_on_web,
    ),
    Tool(
        name="update_venue_contact",
        description=(
            "Write venue contact details to the Gig Booking Worksheet xlsx. "
            "Updates an existing row if the venue is already in the sheet, "
            "otherwise appends a new row. Use this AFTER lookup_venue_email_on_web "
            "finds a booker email for a venue that wasn't in the xlsx — saves it "
            "so next session it's a lookup hit. Returns {updated: True, action: "
            "'updated' or 'created', row: N}."
        ),
        parameters={
            "type": "object",
            "properties": {
                "venue_name": {
                    "type": "string",
                    "description": "Venue name (matches an existing row, or creates a new row if not present).",
                },
                "email": {
                    "type": "string",
                    "description": "Booker email address to save.",
                },
                "contact": {
                    "type": "string",
                    "description": "Optional: booker name.",
                },
                "phone": {
                    "type": "string",
                    "description": "Optional: phone number.",
                },
                "location": {
                    "type": "string",
                    "description": "Optional: city/area (e.g. 'Catawba, VA').",
                },
                "notes": {
                    "type": "string",
                    "description": (
                        "Optional: short note about where the email came from "
                        "(e.g. 'derived from website 2026-05-18'). Will be "
                        "appended to existing notes rather than overwriting."
                    ),
                },
            },
            "required": ["venue_name"],
        },
        handler=update_venue_contact,
    ),
]
